from openpyxl import load_workbook #para arquivos excel
import webbrowser #para abrir o navegador
from urllib.parse import quote #para formatar a URL
import sounddevice as sd # para reconhecimento de voz
import speech_recognition as sr
from langchain_community.llms import Ollama #para LLM

#⠄⠂⠁⠁⠂⠄⠄⠂⠁⠁⠂⠄⠄⠂⠁⠁⠂⠄⠄⠂⠁⠁⠂⠄⠄⠂⠁⠁⠂⠄⠄⠂⠁⠁⠂⠄⠄⠂⠁⠁⠂⠄⠄⠂⠁⠁⠂⠄⠄⠂⠁⠁⠂⠄⠄⠂⠁⠁⠂⠄⠄⠂⠁⠁⠂⠄⠄⠂⠁⠁⠂⠄⠄⠂⠁⠁⠂⠄⠄⠂⠁⠁⠂⠄⠄⠂⠁⠁⠂⠄

ARQUIVO = "lista_compras.xlsx"

#⠄⠂⠁⠁⠂⠄⠄⠂⠁⠁⠂⠄⠄⠂⠁⠁⠂⠄⠄⠂⠁⠁⠂⠄⠄⠂⠁⠁⠂⠄⠄⠂⠁⠁⠂⠄⠄⠂⠁⠁⠂⠄⠄⠂⠁⠁⠂⠄⠄⠂⠁⠁⠂⠄⠄⠂⠁⠁⠂⠄⠄⠂⠁⠁⠂⠄⠄⠂⠁⠁⠂⠄⠄⠂⠁⠁⠂⠄⠄⠂⠁⠁⠂⠄⠄⠂⠁⠁⠂⠄

#manipulação da lista de compras
def adicionar_produto(produto):
    produto = produto.strip()

    if not produto or len(produto) < 2:
        print("Produto inválido.")
        return

    if "|" in produto or "por favor" in produto.lower():
        print("Produto suspeito ignorado.")
        return

    wb = load_workbook(ARQUIVO)
    ws = wb.active
    ws.append([produto])
    wb.save(ARQUIVO)

    print(f"Produto '{produto}' adicionado com sucesso!")


def remover_produto(produto):
    wb = load_workbook(ARQUIVO)
    ws = wb.active
    encontrado = False
    for row in ws.iter_rows(min_row=2):  # pula o cabeçalho
        if row[0].value.lower() == produto.lower():
            ws.delete_rows(row[0].row)
            encontrado = True
            break
    wb.save(ARQUIVO)
    if encontrado:
        print(f"Produto '{produto}' removido da lista!")
    else:
        print(f"Produto '{produto}' não encontrado na lista.")

def listar_produtos():
    wb = load_workbook(ARQUIVO)
    ws = wb.active
    produtos = [row[0].value for row in ws.iter_rows(min_row=2) if row[0].value]
    if produtos:
        print("Lista de compras:")
        for p in produtos:
            print("-", p)
    else:
        print("A lista de compras está vazia.")

#⠄⠂⠁⠁⠂⠄⠄⠂⠁⠁⠂⠄⠄⠂⠁⠁⠂⠄⠄⠂⠁⠁⠂⠄⠄⠂⠁⠁⠂⠄⠄⠂⠁⠁⠂⠄⠄⠂⠁⠁⠂⠄⠄⠂⠁⠁⠂⠄⠄⠂⠁⠁⠂⠄⠄⠂⠁⠁⠂⠄⠄⠂⠁⠁⠂⠄⠄⠂⠁⠁⠂⠄⠄⠂⠁⠁⠂⠄⠄⠂⠁⠁⠂⠄⠄⠂⠁⠁⠂⠄

#para enviar a lista pelo whatsapp

def enviar_wpp():
    wb = load_workbook(ARQUIVO)
    ws = wb.active
    
    produtos = []
    for row in ws.iter_rows(min_row=2):
        valor = row[0].value
        if valor and not str(valor).lower().startswith("aguarde"):
            valor_limpo = str(valor).strip()
            if valor_limpo not in produtos:
                produtos.append(valor_limpo)
    
    if not produtos:
        print("A lista de compras está vazia. Nada há para enviar.")
        return
    
    mensagem = "Lista de compras: " + ", ".join(produtos)
    url = f"https://wa.me/?text={quote(mensagem)}"
    webbrowser.open(url)
    print("Abrindo WhatsApp para enviar a lista...")



#⠄⠂⠁⠁⠂⠄⠄⠂⠁⠁⠂⠄⠄⠂⠁⠁⠂⠄⠄⠂⠁⠁⠂⠄⠄⠂⠁⠁⠂⠄⠄⠂⠁⠁⠂⠄⠄⠂⠁⠁⠂⠄⠄⠂⠁⠁⠂⠄⠄⠂⠁⠁⠂⠄⠄⠂⠁⠁⠂⠄⠄⠂⠁⠁⠂⠄⠄⠂⠁⠁⠂⠄⠄⠂⠁⠁⠂⠄⠄⠂⠁⠁⠂⠄⠄⠂⠁⠁⠂

# Reconhecimento de voz
import speech_recognition as sr

def ouvir_comando():
    r = sr.Recognizer()
    with sr.Microphone() as source:
        print("Diga um comando: ")
        audio = r.listen(source)
    try:
        texto = r.recognize_google(audio, language="pt-BR")
        print(f"Você disse: {texto}")
        return texto
    except sr.UnknownValueError:
        print("Não entendi o que você disse.")
        return ""
    except sr.RequestError:
        print("Erro no serviço de reconhecimento.")
        return ""


    
#⠄⠂⠁⠁⠂⠄⠄⠂⠁⠁⠂⠄⠄⠂⠁⠁⠂⠄⠄⠂⠁⠁⠂⠄⠄⠂⠁⠁⠂⠄⠄⠂⠁⠁⠂⠄⠄⠂⠁⠁⠂⠄⠄⠂⠁⠁⠂⠄⠄⠂⠁⠁⠂⠄⠄⠂⠁⠁⠂⠄⠄⠂⠁⠁⠂⠄⠄⠂⠁⠁⠂⠄⠄⠂⠁⠁⠂⠄⠄⠂⠁⠁⠂⠄⠄⠂⠁⠁⠂

# Interpretar comandos com LLM

llm = Ollama(model="mistral")

from langchain_ollama import OllamaLLM
import unicodedata

# Inicializa o LLM Mistral
llm = OllamaLLM(model="mistral")

# Função para normalizar texto (minúsculas, remove acentos e espaços extras)
def normalizar(texto):
    texto = texto.lower().strip()
    texto = unicodedata.normalize('NFKD', texto).encode('ASCII', 'ignore').decode('ASCII')
    return texto

# Função para interpretar o comando usando o LLM
def interpretar_comando(texto):
    prompt = f"""
    Você é um sistema de classificação de comandos.
    Responda SOMENTE no formato:

    acao|produto

    Sem frases.
    Sem explicações.
    Sem quebras de linha.
    Sem texto adicional.

    Ações possíveis: adicionar, remover, listar, enviar.

    Se não houver produto, deixe vazio após o |.

    Comando: {texto}
    """

    resposta = llm.invoke(prompt)

    resposta = resposta.strip().lower()

    # pega apenas a primeira linha (corta explicações)
    resposta = resposta.split("\n")[0]

    # remove aspas
    resposta = resposta.replace('"', '').replace("'", "")

    print("LLM retornou:", resposta)
    return resposta

#⠄⠂⠁⠁⠂⠄⠄⠂⠁⠁⠂⠄⠄⠂⠁⠁⠂⠄⠄⠂⠁⠁⠂⠄⠄⠂⠁⠁⠂⠄⠄⠂⠁⠁⠂⠄⠄⠂⠁⠁⠂⠄⠄⠂⠁⠁⠂⠄⠄⠂⠁⠁⠂⠄⠄⠂⠁⠁⠂⠄⠄⠂⠁⠁⠂⠄⠄⠂⠁⠁⠂⠄⠄⠂⠁⠁⠂⠄⠄⠂⠁⠁⠂⠄⠄⠂⠁⠁⠂

# Executar os comandos 

def executar_comando(resposta):
    if "|" in resposta:
        acao, produto = resposta.split("|", 1)
        acao = acao.strip()
        produto = produto.strip()
    else:
        acao, produto = resposta, ""
    
    if acao == "adicionar" and produto:
        adicionar_produto(produto)
    elif acao == "remover" and produto:
        remover_produto(produto)
    elif acao == "listar":
        listar_produtos()
    elif acao == "enviar":
        enviar_wpp()
    else:
        print("Comando inválido, tente novamente.")

#⠄⠂⠁⠁⠂⠄⠄⠂⠁⠁⠂⠄⠄⠂⠁⠁⠂⠄⠄⠂⠁⠁⠂⠄⠄⠂⠁⠁⠂⠄⠄⠂⠁⠁⠂⠄⠄⠂⠁⠁⠂⠄⠄⠂⠁⠁⠂⠄⠄⠂⠁⠁⠂⠄⠄⠂⠁⠁⠂⠄⠄⠂⠁⠁⠂⠄⠄⠂⠁⠁⠂⠄⠄⠂⠁⠁⠂⠄⠄⠂⠁⠁⠂⠄⠄⠂⠁⠁⠂

# Loop principal
while True:
    comando = ouvir_comando()
    
    if comando.lower() in ["sair", "encerrar", "fim", "fechar", "finalizar", "deu", "chega"]:
        print("O agente está encerrando. Até mais!")
        break

    if comando:
        resposta = interpretar_comando(comando)
        executar_comando(resposta)


